#!/usr/bin/env python3
"""
Gera o pacote importável do Power Automate (UAT_Extractor_Flow.zip)
e o Excel modelo com Tabela nomeada (UAT_Resultados_Modelo.xlsx).

Executar:  python build_package.py
"""

import json
import uuid
import zipfile
from datetime import datetime, timezone
from pathlib import Path

FLOW_GUID    = "06583cda-9e3b-47f8-ac1d-1c68390f38cb"
CONN_SP_GUID = "6c373451-574f-4b36-b446-45fb8ba97785"
CONN_XL_GUID = "05ef8d86-f09f-42ba-ace8-52a274c3c11e"
GROUP_GUID   = "0451d8d4-7cc5-40a2-9fa5-62e93283a717"

FLOW_NAME  = "UAT - Extrair MS Lists para Excel"
MAX_EVID   = 5
NOW        = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.%f")[:-3] + "Z"

SP_API = "/providers/Microsoft.PowerApps/apis/shared_sharepointonline"
XL_API = "/providers/Microsoft.PowerApps/apis/shared_excelonlinebusiness"

ID_EXPR = "items('Para_cada_item')?['ID']"
PASTA_ITEM = (
    "concat(variables('varBiblioteca'), '/Evidencias/Item_', "
    + ID_EXPR + ")"
)


def _evidencia_expr(i: int) -> str:
    """Expressão da coluna 'Evidência i': =IMAGE(url) se o anexo i existir."""
    return (
        f"@{{if(greater(length(outputs('Obter_anexos')?['body']), {i}), "
        f"concat('=IMAGE(\"', variables('varSiteUrl'), '/', "
        f"variables('varBiblioteca'), '/Evidencias/Item_', {ID_EXPR}, '/', "
        f"outputs('Obter_anexos')?['body'][{i}]?['DisplayName'], '\")'), '')}}"
    )


def build_definition() -> dict:
    add_row_columns = {
        "ID":        f"@{{{ID_EXPR}}}",
        "Cenário":   "@{items('Para_cada_item')?['Title']}",
        "Passo":     "@{items('Para_cada_item')?['Passo']}",
        "Resultado": "@{items('Para_cada_item')?['Resultado']}",
        "Link": (
            "@{if(greater(length(outputs('Obter_anexos')?['body']), 0), "
            "concat('=HYPERLINK(\"', variables('varSiteUrl'), '/', "
            "variables('varBiblioteca'), '/Evidencias/Item_', "
            + ID_EXPR + ", '\",\"📷 Abrir pasta\")'), '')}"
        ),
    }
    for i in range(MAX_EVID):
        add_row_columns[f"Evidência {i+1}"] = _evidencia_expr(i)

    definition = {
        "$schema": "https://schema.management.azure.com/providers/Microsoft.Logic/schemas/2016-06-01/workflowdefinition.json#",
        "contentVersion": "1.0.0.0",
        "parameters": {
            "$connections": {"defaultValue": {}, "type": "Object"},
            "$authentication": {"defaultValue": {}, "type": "SecureObject"},
        },
        "triggers": {
            "manual": {
                "type": "Request",
                "kind": "Button",
                "inputs": {"schema": {"type": "object", "properties": {}, "required": []}},
            }
        },
        "actions": {
            # ------------------------------------------------------------
            # ⚙️ CONFIGURAÇÃO — os 3 únicos valores a editar após importar
            # ------------------------------------------------------------
            "varSiteUrl": {
                "type": "InitializeVariable",
                "runAfter": {},
                "inputs": {"variables": [{
                    "name": "varSiteUrl", "type": "string",
                    "value": "https://EMPRESA.sharepoint.com/sites/SEUSITE",
                }]},
            },
            "varNomeLista": {
                "type": "InitializeVariable",
                "runAfter": {"varSiteUrl": ["Succeeded"]},
                "inputs": {"variables": [{
                    "name": "varNomeLista", "type": "string",
                    "value": "UAT - Cenarios e Passos",
                }]},
            },
            "varBiblioteca": {
                "type": "InitializeVariable",
                "runAfter": {"varNomeLista": ["Succeeded"]},
                "inputs": {"variables": [{
                    "name": "varBiblioteca", "type": "string",
                    "value": "Documentos Compartilhados/UAT",
                }]},
            },
            # ------------------------------------------------------------
            "Obter_itens": {
                "type": "OpenApiConnection",
                "runAfter": {"varBiblioteca": ["Succeeded"]},
                "inputs": {
                    "host": {
                        "connectionName": "shared_sharepointonline",
                        "operationId": "GetItems",
                        "apiId": SP_API,
                    },
                    "parameters": {
                        "dataset": "@variables('varSiteUrl')",
                        "table": "@variables('varNomeLista')",
                    },
                    "authentication": "@parameters('$authentication')",
                },
            },
            "Para_cada_item": {
                "type": "Foreach",
                "runAfter": {"Obter_itens": ["Succeeded"]},
                "foreach": "@outputs('Obter_itens')?['body/value']",
                "runtimeConfiguration": {"concurrency": {"repetitions": 1}},
                "actions": {
                    "Obter_anexos": {
                        "type": "OpenApiConnection",
                        "runAfter": {},
                        "inputs": {
                            "host": {
                                "connectionName": "shared_sharepointonline",
                                "operationId": "GetAttachments",
                                "apiId": SP_API,
                            },
                            "parameters": {
                                "dataset": "@variables('varSiteUrl')",
                                "table": "@variables('varNomeLista')",
                                "id": f"@{ID_EXPR}",
                            },
                            "authentication": "@parameters('$authentication')",
                        },
                    },
                    "Para_cada_anexo": {
                        "type": "Foreach",
                        "runAfter": {"Obter_anexos": ["Succeeded"]},
                        "foreach": "@outputs('Obter_anexos')?['body']",
                        "actions": {
                            "Conteudo_do_anexo": {
                                "type": "OpenApiConnection",
                                "runAfter": {},
                                "inputs": {
                                    "host": {
                                        "connectionName": "shared_sharepointonline",
                                        "operationId": "GetAttachmentContent",
                                        "apiId": SP_API,
                                    },
                                    "parameters": {
                                        "dataset": "@variables('varSiteUrl')",
                                        "table": "@variables('varNomeLista')",
                                        "id": f"@{ID_EXPR}",
                                        "attachmentId": "@items('Para_cada_anexo')?['Id']",
                                    },
                                    "authentication": "@parameters('$authentication')",
                                },
                            },
                            "Salvar_imagem": {
                                "type": "OpenApiConnection",
                                "runAfter": {"Conteudo_do_anexo": ["Succeeded"]},
                                "inputs": {
                                    "host": {
                                        "connectionName": "shared_sharepointonline",
                                        "operationId": "CreateFile",
                                        "apiId": SP_API,
                                    },
                                    "parameters": {
                                        "dataset": "@variables('varSiteUrl')",
                                        "folderPath": f"@{PASTA_ITEM}",
                                        "name": "@items('Para_cada_anexo')?['DisplayName']",
                                        "body": "@outputs('Conteudo_do_anexo')?['body']",
                                    },
                                    "authentication": "@parameters('$authentication')",
                                },
                            },
                        },
                    },
                    "Adicionar_linha_no_Excel": {
                        "type": "OpenApiConnection",
                        "runAfter": {"Para_cada_anexo": ["Succeeded"]},
                        "inputs": {
                            "host": {
                                "connectionName": "shared_excelonlinebusiness",
                                "operationId": "AddRowV2",
                                "apiId": XL_API,
                            },
                            "parameters": {
                                # Estes 4 campos são reapontados no designer após importar
                                # (selecionar o arquivo UAT_Resultados_Modelo.xlsx e a TabelaUAT)
                                "source": "@variables('varSiteUrl')",
                                "drive": "SELECIONAR_NO_DESIGNER",
                                "file": "SELECIONAR_NO_DESIGNER",
                                "table": "TabelaUAT",
                                "item": add_row_columns,
                            },
                            "authentication": "@parameters('$authentication')",
                        },
                    },
                },
            },
        },
        "outputs": {},
    }
    return definition


def build_package(out_zip: str):
    definition = build_definition()

    flow_manifest = {
        "schemaVersion": "1.0.0.0",
        "properties": {
            "displayName": FLOW_NAME,
            "definition": definition,
            "connectionReferences": {
                "shared_sharepointonline": {
                    "connectionName": "shared-sharepointonline",
                    "source": "Invoker",
                    "id": SP_API,
                    "tier": "NotSpecified",
                },
                "shared_excelonlinebusiness": {
                    "connectionName": "shared-excelonlinebusiness",
                    "source": "Invoker",
                    "id": XL_API,
                    "tier": "NotSpecified",
                },
            },
        },
    }

    root_manifest = {
        "schema": "1.0",
        "details": {
            "displayName": FLOW_NAME,
            "description": "Extrai cenários/passos de uma MS List, salva as imagens anexadas em pastas por ID e monta a planilha Excel com =IMAGE() e link por item.",
            "createdTime": NOW,
            "packageTelemetryId": GROUP_GUID,
            "creator": "N/A",
            "sourceEnvironment": "",
        },
        "resources": {
            FLOW_GUID: {
                "id": f"/providers/Microsoft.Flow/flows/{FLOW_GUID}",
                "name": FLOW_GUID,
                "type": "Microsoft.Flow/flows",
                "creationType": "Existing, New",
                "details": {"displayName": FLOW_NAME},
                "configurableBy": "User",
                "hierarchy": "Root",
                "dependsOn": [CONN_SP_GUID, CONN_XL_GUID],
            },
            CONN_SP_GUID: {
                "id": SP_API,
                "name": "shared_sharepointonline",
                "type": "Microsoft.PowerApps/apis",
                "suggestedCreationType": "Existing",
                "creationType": "Existing",
                "details": {
                    "displayName": "SharePoint",
                    "iconUri": "https://connectoricons-prod.azureedge.net/sharepointonline/icon.png",
                },
                "configurableBy": "User",
                "hierarchy": "Child",
                "dependsOn": [],
            },
            CONN_XL_GUID: {
                "id": XL_API,
                "name": "shared_excelonlinebusiness",
                "type": "Microsoft.PowerApps/apis",
                "suggestedCreationType": "Existing",
                "creationType": "Existing",
                "details": {
                    "displayName": "Excel Online (Business)",
                    "iconUri": "https://connectoricons-prod.azureedge.net/excelonlinebusiness/icon.png",
                },
                "configurableBy": "User",
                "hierarchy": "Child",
                "dependsOn": [],
            },
        },
    }

    api_map = {
        "shared_sharepointonline": SP_API,
        "shared_excelonlinebusiness": XL_API,
    }

    with zipfile.ZipFile(out_zip, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("manifest.json", json.dumps(root_manifest, ensure_ascii=False, indent=2))
        base = f"Microsoft.Flow/flows/{FLOW_GUID}"
        z.writestr(f"{base}/definition.json", json.dumps(flow_manifest, ensure_ascii=False, indent=2))
        z.writestr(f"{base}/apisMap.json", json.dumps(api_map, ensure_ascii=False, indent=2))

    print(f"✅  Pacote gerado: {out_zip}")


def build_excel(out_xlsx: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    wb = Workbook()
    ws = wb.active
    ws.title = "UAT"
    ws.sheet_view.showGridLines = False

    headers = (["ID", "Cenário", "Passo", "Resultado"]
               + [f"Evidência {i+1}" for i in range(MAX_EVID)]
               + ["Link"])
    widths = [10, 30, 30, 14] + [30] * MAX_EVID + [16]

    lado = Side(style="thin", color="BFBFBF")
    borda = Border(left=lado, right=lado, top=lado, bottom=lado)

    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F3864")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = borda
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 26

    last_col = get_column_letter(len(headers))
    tabela = Table(displayName="TabelaUAT", ref=f"A1:{last_col}2")
    tabela.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(tabela)

    wb.save(out_xlsx)
    print(f"✅  Excel modelo gerado: {out_xlsx}")


if __name__ == "__main__":
    aqui = Path(__file__).parent
    build_package(str(aqui / "UAT_Extractor_Flow.zip"))
    build_excel(str(aqui / "UAT_Resultados_Modelo.xlsx"))
