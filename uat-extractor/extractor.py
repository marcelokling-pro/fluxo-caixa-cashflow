#!/usr/bin/env python3
"""
UAT MS Lists Extractor
Extrai cenários, passos e imagens de uma MS List e gera um Excel formatado.
"""

import io
import os
import sys
import re
import json
import time
import tempfile
import requests
from datetime import datetime, date
from pathlib import Path

try:
    import msal
except ImportError:
    sys.exit("❌  Dependência ausente. Execute: pip install -r requirements.txt")

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.series import DataPoint
except ImportError:
    sys.exit("❌  Dependência ausente. Execute: pip install -r requirements.txt")

try:
    from PIL import Image as PILImage
except ImportError:
    PILImage = None  # imagens serão embutidas sem redimensionamento

import config

# ---------------------------------------------------------------------------
# Paleta de cores
# ---------------------------------------------------------------------------
COR_HEADER_BG   = "1F3864"   # azul escuro
COR_HEADER_FG   = "FFFFFF"
COR_APROVADO    = "C6EFCE"   # verde claro
COR_APROVADO_FG = "276221"
COR_REPROVADO   = "FFC7CE"   # vermelho claro
COR_REPROVADO_FG= "9C0006"
COR_BLOQUEADO   = "FFEB9C"   # amarelo
COR_BLOQUEADO_FG= "9C6500"
COR_NAO_EXEC    = "D9D9D9"   # cinza
COR_NAO_EXEC_FG = "595959"
COR_LINHA_ALT   = "EEF3FB"   # azul bem claro (linhas alternadas)
COR_TITULO      = "2E75B6"   # azul médio (título da planilha)
COR_SECAO       = "D6E4F0"   # azul claro (cabeçalho de cenário)

SCOPES = ["https://graph.microsoft.com/.default"]
GRAPH  = "https://graph.microsoft.com/v1.0"


# ---------------------------------------------------------------------------
# Autenticação (Device Code Flow — sem segredo de app)
# ---------------------------------------------------------------------------

def autenticar() -> str:
    app = msal.PublicClientApplication(
        client_id=config.CLIENT_ID,
        authority=f"https://login.microsoftonline.com/{config.TENANT_ID}",
    )
    # tenta token em cache primeiro
    contas = app.get_accounts()
    if contas:
        resultado = app.acquire_token_silent(SCOPES, account=contas[0])
        if resultado and "access_token" in resultado:
            print("✅  Token obtido do cache.")
            return resultado["access_token"]

    flow = app.initiate_device_flow(scopes=SCOPES)
    if "user_code" not in flow:
        sys.exit(f"❌  Falha ao iniciar device flow: {flow}")

    print("\n" + "="*60)
    print("🔐  AUTENTICAÇÃO NECESSÁRIA")
    print("="*60)
    print(flow["message"])
    print("="*60 + "\n")

    resultado = app.acquire_token_by_device_flow(flow)
    if "access_token" not in resultado:
        sys.exit(f"❌  Falha na autenticação: {resultado.get('error_description')}")

    print("✅  Autenticado com sucesso.\n")
    return resultado["access_token"]


# ---------------------------------------------------------------------------
# Funções de acesso ao Graph API
# ---------------------------------------------------------------------------

def _headers(token: str) -> dict:
    return {"Authorization": f"Bearer {token}", "Accept": "application/json"}


def _get(token: str, url: str, params: dict = None) -> dict:
    r = requests.get(url, headers=_headers(token), params=params, timeout=30)
    r.raise_for_status()
    return r.json()


def _get_paginado(token: str, url: str, params: dict = None) -> list:
    """Itera sobre páginas @odata.nextLink automaticamente."""
    itens = []
    while url:
        dados = _get(token, url, params)
        itens.extend(dados.get("value", []))
        url = dados.get("@odata.nextLink")
        params = None  # nextLink já contém os params
    return itens


def obter_site_id(token: str) -> str:
    """Converte SHAREPOINT_URL em site-id do Graph."""
    url = config.SHAREPOINT_URL.rstrip("/")
    # formato: https://tenant.sharepoint.com/sites/nome
    m = re.match(r"https://([^/]+)(/.+)", url)
    if not m:
        sys.exit("❌  SHAREPOINT_URL inválida. Formato esperado: https://empresa.sharepoint.com/sites/site")
    host, path = m.groups()
    endpoint = f"{GRAPH}/sites/{host}:{path}"
    dados = _get(token, endpoint)
    return dados["id"]


def obter_list_id(token: str, site_id: str) -> str:
    listas = _get_paginado(token, f"{GRAPH}/sites/{site_id}/lists",
                           {"$select": "id,name,displayName"})
    for l in listas:
        if l.get("displayName") == config.LIST_NAME or l.get("name") == config.LIST_NAME:
            return l["id"]
    nomes = [l.get("displayName", l.get("name")) for l in listas]
    sys.exit(f"❌  List '{config.LIST_NAME}' não encontrada. Listas disponíveis: {nomes}")


def obter_itens(token: str, site_id: str, list_id: str) -> list:
    print("⬇️   Baixando itens da MS List...")
    campos = ",".join(config.FIELD_MAP.values())
    url = f"{GRAPH}/sites/{site_id}/lists/{list_id}/items"
    itens = _get_paginado(token, url, {
        "$expand": "fields($select=" + campos + ")",
        "$top": 500,
    })
    print(f"     {len(itens)} itens obtidos.")
    return itens


def obter_anexos(token: str, site_id: str, list_id: str, item_id: str) -> list:
    """Retorna lista de {nome, bytes} para cada anexo de imagem do item."""
    url = f"{GRAPH}/sites/{site_id}/lists/{list_id}/items/{item_id}/attachments"
    try:
        dados = _get(token, url)
    except requests.HTTPError:
        return []

    anexos = []
    for a in dados.get("value", []):
        nome = a.get("name", "")
        if not _eh_imagem(nome):
            continue
        content_url = a.get("contentUrl") or \
            f"{GRAPH}/sites/{site_id}/lists/{list_id}/items/{item_id}/attachments/{a['id']}/$value"
        try:
            r = requests.get(content_url, headers=_headers(token), timeout=60)
            r.raise_for_status()
            anexos.append({"nome": nome, "dados": r.content})
        except Exception as e:
            print(f"     ⚠️   Não foi possível baixar '{nome}': {e}")
    return anexos


def _eh_imagem(nome: str) -> bool:
    return Path(nome).suffix.lower() in {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp", ".tiff"}


# ---------------------------------------------------------------------------
# Normalização dos dados
# ---------------------------------------------------------------------------

def normalizar(itens: list) -> list:
    """Converte lista de itens Graph → lista de dicts com campos padronizados."""
    fm = config.FIELD_MAP
    resultado = []
    for item in itens:
        f = item.get("fields", {})
        linha = {k: f.get(v, "") for k, v in fm.items()}
        linha["_item_id"] = item["id"]
        # garante que data seja string legível
        if linha.get("data_execucao"):
            try:
                dt = datetime.fromisoformat(linha["data_execucao"].replace("Z", "+00:00"))
                linha["data_execucao"] = dt.strftime("%d/%m/%Y")
            except Exception:
                pass
        resultado.append(linha)

    # ordena por cenário e depois por passo
    def chave(r):
        try:
            c = int(str(r.get("id_cenario", 0)).split(".")[0])
        except Exception:
            c = 0
        try:
            p = int(str(r.get("id_passo", 0)).split(".")[0])
        except Exception:
            p = 0
        return (c, p)

    resultado.sort(key=chave)
    return resultado


# ---------------------------------------------------------------------------
# Redimensionamento de imagem
# ---------------------------------------------------------------------------

def _redimensionar(img_bytes: bytes) -> bytes:
    if PILImage is None:
        return img_bytes
    try:
        img = PILImage.open(io.BytesIO(img_bytes))
        img.thumbnail((config.IMAGE_MAX_WIDTH, config.IMAGE_MAX_HEIGHT), PILImage.LANCZOS)
        buf = io.BytesIO()
        fmt = img.format or "PNG"
        if fmt == "JPEG" or fmt == "JPG":
            img.save(buf, format="JPEG", quality=85)
        else:
            img.save(buf, format="PNG")
        return buf.getvalue()
    except Exception:
        return img_bytes


# ---------------------------------------------------------------------------
# Geração do Excel
# ---------------------------------------------------------------------------

COLUNAS = [
    ("ID\nCenário",       "id_cenario",        8),
    ("Cenário",           "nome_cenario",      30),
    ("Passo\nNº",         "id_passo",           7),
    ("Descrição do Passo","descricao_passo",   40),
    ("Resultado\nEsperado","resultado_esperado",35),
    ("Resultado\nObtido", "resultado_obtido",  35),
    ("Status",            "status",            14),
    ("Responsável",       "responsavel",       18),
    ("Data\nExecução",    "data_execucao",     13),
    ("Ambiente",          "ambiente",          12),
    ("Versão",            "versao",            10),
    ("Observações",       "observacoes",       30),
    ("Evidências",        "evidencias",        35),  # coluna sintética de imagens
]


def _borda_fina():
    lado = Side(style="thin", color="BFBFBF")
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def _fill(hex_cor: str):
    return PatternFill("solid", fgColor=hex_cor)


def _cor_status(status: str):
    s = (status or "").strip().lower()
    if "aprovad" in s or "pass" in s or "ok" in s:
        return COR_APROVADO, COR_APROVADO_FG
    if "reprovad" in s or "fail" in s or "nok" in s:
        return COR_REPROVADO, COR_REPROVADO_FG
    if "bloqueado" in s or "block" in s:
        return COR_BLOQUEADO, COR_BLOQUEADO_FG
    return COR_NAO_EXEC, COR_NAO_EXEC_FG


def gerar_excel(linhas: list, anexos_por_item: dict, output_path: str):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove Sheet padrão

    _sheet_capa(wb, linhas)
    _sheet_detalhes(wb, linhas, anexos_por_item)
    if config.SHEET_SUMMARY:
        _sheet_resumo(wb, linhas)

    wb.save(output_path)
    print(f"\n✅  Excel salvo em: {output_path}")


# -- Aba CAPA ----------------------------------------------------------------

def _sheet_capa(wb, linhas: list):
    ws = wb.create_sheet("Capa")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 25

    # fundo azul escuro nas primeiras linhas
    for row in range(1, 12):
        for col in range(1, 10):
            ws.cell(row=row, column=col).fill = _fill(COR_HEADER_BG)

    def capa_cell(row, col, value, size=11, bold=False, color="FFFFFF", wrap=False):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name="Calibri", size=size, bold=bold, color=color)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=wrap)
        return c

    ws.row_dimensions[2].height = 40
    capa_cell(2, 2, "UAT — Relatório de Execução", size=22, bold=True)

    ws.row_dimensions[4].height = 20
    capa_cell(4, 2, f"Data de extração:  {date.today().strftime('%d/%m/%Y')}", size=12)

    ws.row_dimensions[5].height = 20
    capa_cell(5, 2, f"Fonte:  {config.SHAREPOINT_URL}", size=11)

    ws.row_dimensions[6].height = 20
    capa_cell(6, 2, f"MS List:  {config.LIST_NAME}", size=11)

    # contadores rápidos
    total   = len(linhas)
    aprov   = sum(1 for l in linhas if "aprovad" in str(l.get("status","")).lower() or "pass" in str(l.get("status","")).lower())
    reprov  = sum(1 for l in linhas if "reprovad" in str(l.get("status","")).lower() or "fail" in str(l.get("status","")).lower())
    bloq    = sum(1 for l in linhas if "bloqueado" in str(l.get("status","")).lower())
    n_exec  = total - aprov - reprov - bloq

    pct = lambda n: f"{round(n/total*100)}%" if total else "—"

    ws.row_dimensions[8].height = 18
    capa_cell(8, 2, "Resumo geral", size=13, bold=True)

    dados_resumo = [
        ("Total de passos",      total,  "FFFFFF"),
        ("✔  Aprovados",         aprov,  COR_APROVADO),
        ("✘  Reprovados",        reprov, COR_REPROVADO),
        ("⚠  Bloqueados",        bloq,   COR_BLOQUEADO),
        ("○  Não executados",    n_exec, COR_NAO_EXEC),
    ]
    for i, (label, valor, cor) in enumerate(dados_resumo, start=9):
        ws.row_dimensions[i].height = 20
        c_label = ws.cell(row=i, column=2, value=label)
        c_label.font = Font(name="Calibri", size=11, color="FFFFFF")
        c_label.alignment = Alignment(vertical="center")
        c_val = ws.cell(row=i, column=3, value=valor)
        c_val.font = Font(name="Calibri", size=11, bold=True, color="111111" if cor != "FFFFFF" else "FFFFFF")
        c_val.fill = _fill(cor)
        c_val.alignment = Alignment(horizontal="center", vertical="center")
        c_val.border = _borda_fina()


# -- Aba DETALHES ------------------------------------------------------------

def _sheet_detalhes(wb, linhas: list, anexos_por_item: dict):
    ws = wb.create_sheet("Detalhes UAT")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    # cabeçalho duplo (linha 1: título, linha 2: colunas)
    ws.merge_cells("A1:M1")
    titulo = ws["A1"]
    titulo.value = f"UAT — Detalhamento de Execução   |   {config.LIST_NAME}   |   {date.today().strftime('%d/%m/%Y')}"
    titulo.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    titulo.fill = _fill(COR_TITULO)
    titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # cabeçalho das colunas (linha 2)
    ws.row_dimensions[2].height = 36
    for col_idx, (header, _, width) in enumerate(COLUNAS, start=1):
        c = ws.cell(row=2, column=col_idx, value=header)
        c.font = Font(name="Calibri", size=10, bold=True, color=COR_HEADER_FG)
        c.fill = _fill(COR_HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = _borda_fina()
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # linhas de dados
    linha_excel = 3
    cenario_atual = None

    for idx, linha in enumerate(linhas):
        is_alt = (idx % 2 == 1)
        cenario = linha.get("id_cenario", "")
        novo_cenario = (cenario != cenario_atual)
        cenario_atual = cenario

        # linha separadora de cenário
        if novo_cenario and idx > 0:
            for col in range(1, len(COLUNAS) + 1):
                ws.cell(row=linha_excel, column=col).fill = _fill("DDEEFF")
            ws.row_dimensions[linha_excel].height = 4
            linha_excel += 1

        # calcula a altura necessária para imagens
        anexos = anexos_por_item.get(linha["_item_id"], [])
        n_imgs = len(anexos)
        altura_linha = max(20, n_imgs * (config.IMAGE_MAX_HEIGHT + 6)) if (config.EMBED_IMAGES and n_imgs) else 20
        ws.row_dimensions[linha_excel].height = altura_linha

        for col_idx, (_, campo, _) in enumerate(COLUNAS, start=1):
            if campo == "evidencias":
                continue  # tratado separadamente abaixo
            valor = linha.get(campo, "")
            c = ws.cell(row=linha_excel, column=col_idx, value=valor)
            c.font = Font(name="Calibri", size=10)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = _borda_fina()

            status = linha.get("status", "")
            if campo == "status" and status:
                bg, fg = _cor_status(status)
                c.fill = _fill(bg)
                c.font = Font(name="Calibri", size=10, bold=True, color=fg)
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif is_alt:
                c.fill = _fill(COR_LINHA_ALT)

        # coluna de evidências (imagens ou nomes)
        ev_col = len(COLUNAS)  # última coluna
        ev_cell = ws.cell(row=linha_excel, column=ev_col)
        ev_cell.border = _borda_fina()
        if is_alt:
            ev_cell.fill = _fill(COR_LINHA_ALT)

        if not anexos:
            ev_cell.value = "—"
            ev_cell.font = Font(name="Calibri", size=10, color="999999")
            ev_cell.alignment = Alignment(horizontal="center", vertical="center")
        elif config.EMBED_IMAGES:
            # insere imagens empilhadas na célula
            cell_ref = f"{get_column_letter(ev_col)}{linha_excel}"
            offset_y = 2
            for anx in anexos:
                img_bytes = _redimensionar(anx["dados"])
                try:
                    img_obj = XLImage(io.BytesIO(img_bytes))
                    img_obj.anchor = cell_ref
                    # ancoragem por posição absoluta não é trivial no openpyxl;
                    # empilhamos via AnchorOffset quando disponível
                    ws.add_image(img_obj)
                    offset_y += config.IMAGE_MAX_HEIGHT + 6
                except Exception as e:
                    ev_cell.value = (ev_cell.value or "") + f"[{anx['nome']}] "
        else:
            ev_cell.value = "\n".join(a["nome"] for a in anexos)
            ev_cell.alignment = Alignment(vertical="top", wrap_text=True)

        linha_excel += 1

    # auto-filtro nas colunas de dados
    ws.auto_filter.ref = f"A2:{get_column_letter(len(COLUNAS))}{linha_excel - 1}"


# -- Aba RESUMO --------------------------------------------------------------

def _sheet_resumo(wb, linhas: list):
    ws = wb.create_sheet("Resumo")
    ws.sheet_view.showGridLines = False

    # cabeçalho
    ws.merge_cells("A1:F1")
    h = ws["A1"]
    h.value = "Resumo por Cenário"
    h.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    h.fill = _fill(COR_TITULO)
    h.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = ["ID Cenário", "Nome do Cenário", "Total Passos",
               "Aprovados", "Reprovados", "Bloqueados"]
    widths  = [12, 40, 14, 14, 14, 14]
    for i, (h_txt, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=2, column=i, value=h_txt)
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.fill = _fill(COR_HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _borda_fina()
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[2].height = 30

    # agrupa por cenário
    cenarios: dict = {}
    for l in linhas:
        cid = l.get("id_cenario", "?")
        if cid not in cenarios:
            cenarios[cid] = {"nome": l.get("nome_cenario", ""), "passos": []}
        cenarios[cid]["passos"].append(l.get("status", ""))

    row = 3
    for i, (cid, info) in enumerate(cenarios.items()):
        passos  = info["passos"]
        total   = len(passos)
        aprov   = sum(1 for s in passos if "aprovad" in s.lower() or "pass" in s.lower())
        reprov  = sum(1 for s in passos if "reprovad" in s.lower() or "fail" in s.lower())
        bloq    = sum(1 for s in passos if "bloqueado" in s.lower())
        alt     = i % 2 == 1

        vals = [cid, info["nome"], total, aprov, reprov, bloq]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=row, column=j, value=v)
            c.font = Font(name="Calibri", size=10)
            c.border = _borda_fina()
            c.alignment = Alignment(vertical="center",
                                    horizontal="center" if j != 2 else "left",
                                    wrap_text=True)
            if alt:
                c.fill = _fill(COR_LINHA_ALT)
            # colorir coluna de status
            if j == 4 and aprov == total:
                c.fill = _fill(COR_APROVADO)
                c.font = Font(name="Calibri", size=10, bold=True, color=COR_APROVADO_FG)
            elif j == 5 and reprov > 0:
                c.fill = _fill(COR_REPROVADO)
                c.font = Font(name="Calibri", size=10, bold=True, color=COR_REPROVADO_FG)
            elif j == 6 and bloq > 0:
                c.fill = _fill(COR_BLOQUEADO)
                c.font = Font(name="Calibri", size=10, bold=True, color=COR_BLOQUEADO_FG)
        row += 1

    ws.auto_filter.ref = f"A2:F{row - 1}"


# ---------------------------------------------------------------------------
# Upload para SharePoint
# ---------------------------------------------------------------------------

def fazer_upload(token: str, site_id: str, output_path: str):
    nome_arquivo = Path(output_path).name
    biblioteca   = config.EXCEL_LIBRARY.strip("/")
    url = f"{GRAPH}/sites/{site_id}/drives"
    drives = _get(token, url).get("value", [])

    # usa o primeiro drive (Documents) se não achar pela biblioteca
    drive_id = drives[0]["id"] if drives else None

    upload_url = f"{GRAPH}/sites/{site_id}/drives/{drive_id}/root:/{biblioteca}/{nome_arquivo}:/content"
    with open(output_path, "rb") as f:
        conteudo = f.read()

    headers = {**_headers(token), "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
    r = requests.put(upload_url, headers=headers, data=conteudo, timeout=120)
    if r.status_code in (200, 201):
        link = r.json().get("webUrl", "")
        print(f"☁️   Upload concluído: {link}")
    else:
        print(f"⚠️   Upload retornou {r.status_code}: {r.text[:200]}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("\n" + "="*60)
    print("  UAT MS Lists Extractor")
    print("="*60)

    token    = autenticar()
    site_id  = obter_site_id(token)
    list_id  = obter_list_id(token, site_id)
    itens    = obter_itens(token, site_id, list_id)
    linhas   = normalizar(itens)

    # baixa anexos
    print("\n⬇️   Baixando anexos (imagens)...")
    anexos_por_item: dict = {}
    for i, item in enumerate(itens, start=1):
        iid = item["id"]
        anx = obter_anexos(token, site_id, list_id, iid)
        if anx:
            anexos_por_item[iid] = anx
            print(f"     [{i}/{len(itens)}] item {iid}: {len(anx)} imagem(ns)")
        else:
            sys.stdout.write(f"\r     [{i}/{len(itens)}] verificando...  ")
            sys.stdout.flush()
    print()

    # gera Excel
    data_hoje   = date.today().strftime("%Y-%m-%d")
    output_name = config.OUTPUT_FILENAME.replace("{data}", data_hoje)
    output_path = str(Path(output_name).resolve())

    print("📊  Gerando Excel...")
    gerar_excel(linhas, anexos_por_item, output_path)

    # upload opcional
    resp = input("\n☁️   Deseja fazer upload do Excel para o SharePoint? [s/N] ").strip().lower()
    if resp in ("s", "sim", "y", "yes"):
        fazer_upload(token, site_id, output_path)

    # salva imagens em pasta separada se configurado assim
    if not config.EMBED_IMAGES and any(anexos_por_item.values()):
        pasta_ev = Path(output_path).parent / "evidencias"
        pasta_ev.mkdir(exist_ok=True)
        for iid, anx_list in anexos_por_item.items():
            for anx in anx_list:
                dest = pasta_ev / f"{iid}_{anx['nome']}"
                dest.write_bytes(anx["dados"])
        print(f"🖼️   Evidências salvas em: {pasta_ev}")

    print("\n✅  Concluído.\n")


if __name__ == "__main__":
    main()
