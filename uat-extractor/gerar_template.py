#!/usr/bin/env python3
"""
Gera o arquivo UAT_Template.xlsx com formatação pronta e dados de exemplo.
Use este template para preencher manualmente ou como referência visual.
"""

import io
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COR_HEADER_BG   = "1F3864"
COR_HEADER_FG   = "FFFFFF"
COR_APROVADO    = "C6EFCE"
COR_APROVADO_FG = "276221"
COR_REPROVADO   = "FFC7CE"
COR_REPROVADO_FG= "9C0006"
COR_BLOQUEADO   = "FFEB9C"
COR_BLOQUEADO_FG= "9C6500"
COR_NAO_EXEC    = "D9D9D9"
COR_NAO_EXEC_FG = "595959"
COR_LINHA_ALT   = "EEF3FB"
COR_TITULO      = "2E75B6"

def _fill(hex_cor):
    return PatternFill("solid", fgColor=hex_cor)

def _borda_fina():
    lado = Side(style="thin", color="BFBFBF")
    return Border(left=lado, right=lado, top=lado, bottom=lado)

def _cor_status(status):
    s = status.lower()
    if "aprovad" in s or "pass" in s:
        return COR_APROVADO, COR_APROVADO_FG
    if "reprovad" in s or "fail" in s:
        return COR_REPROVADO, COR_REPROVADO_FG
    if "bloqueado" in s or "block" in s:
        return COR_BLOQUEADO, COR_BLOQUEADO_FG
    return COR_NAO_EXEC, COR_NAO_EXEC_FG

COLUNAS = [
    ("ID\nCenário",         8),
    ("Cenário",            30),
    ("Passo\nNº",           7),
    ("Descrição do Passo", 40),
    ("Resultado\nEsperado",35),
    ("Resultado\nObtido",  35),
    ("Status",             14),
    ("Responsável",        18),
    ("Data\nExecução",     13),
    ("Ambiente",           12),
    ("Versão",             10),
    ("Observações",        30),
    ("Evidências",         35),
]

DADOS_EXEMPLO = [
    # id_cen, nome_cen,               id_pass, descricao,                            esperado,                            obtido,                              status,              resp,           data,         amb,   ver,   obs
    ("C-01", "Login de usuário",       "1",  "Acessar a URL do sistema",            "Tela de login exibida",              "Tela de login exibida",              "Aprovado",          "Ana Silva",    "14/07/2026", "UAT", "2.4", ""),
    ("C-01", "Login de usuário",       "2",  "Inserir credenciais válidas",         "Campo senha aceita entrada",         "Campo senha aceita entrada",         "Aprovado",          "Ana Silva",    "14/07/2026", "UAT", "2.4", ""),
    ("C-01", "Login de usuário",       "3",  "Clicar em Entrar",                   "Redireciona ao dashboard",           "Redireciona ao dashboard",           "Aprovado",          "Ana Silva",    "14/07/2026", "UAT", "2.4", ""),
    ("C-02", "Cadastro de produto",    "1",  "Acessar menu Produtos > Novo",        "Formulário de cadastro aberto",      "Formulário de cadastro aberto",      "Aprovado",          "Carlos Lima",  "14/07/2026", "UAT", "2.4", ""),
    ("C-02", "Cadastro de produto",    "2",  "Preencher nome e preço",              "Campos obrigatórios validados",      "Mensagem de erro ao deixar vazio",   "Reprovado",         "Carlos Lima",  "14/07/2026", "UAT", "2.4", "Validação não disparou"),
    ("C-02", "Cadastro de produto",    "3",  "Clicar em Salvar",                   "Produto salvo e listado",            "Não executado por bloqueio no passo 2","Bloqueado",        "Carlos Lima",  "14/07/2026", "UAT", "2.4", "Aguardando correção"),
    ("C-03", "Relatório mensal",       "1",  "Acessar Relatórios > Mensal",        "Filtros de data exibidos",           "Filtros de data exibidos",           "Aprovado",          "Beatriz Costa","14/07/2026", "UAT", "2.4", ""),
    ("C-03", "Relatório mensal",       "2",  "Selecionar período e exportar PDF",   "Arquivo PDF gerado e baixado",       "",                                   "Não executado",     "Beatriz Costa","",           "UAT", "2.4", "Aguardando ambiente"),
]

def gerar():
    wb = Workbook()
    wb.remove(wb.active)

    # ── Aba Capa ──────────────────────────────────────────────────────────────
    ws_capa = wb.create_sheet("Capa")
    ws_capa.sheet_view.showGridLines = False
    ws_capa.column_dimensions["A"].width = 2
    ws_capa.column_dimensions["B"].width = 50
    ws_capa.column_dimensions["C"].width = 20

    for r in range(1, 14):
        for c in range(1, 9):
            ws_capa.cell(row=r, column=c).fill = _fill(COR_HEADER_BG)

    def c_cell(row, col, val, size=11, bold=False, color="FFFFFF"):
        cell = ws_capa.cell(row=row, column=col, value=val)
        cell.font = Font(name="Calibri", size=size, bold=bold, color=color)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        return cell

    ws_capa.row_dimensions[2].height = 42
    c_cell(2, 2, "UAT — Relatório de Execução", size=22, bold=True)

    ws_capa.row_dimensions[4].height = 20
    c_cell(4, 2, f"Data de extração:  {date.today().strftime('%d/%m/%Y')}", size=12)
    ws_capa.row_dimensions[5].height = 20
    c_cell(5, 2, "Fonte:  SharePoint / MS Lists", size=11)
    ws_capa.row_dimensions[6].height = 20
    c_cell(6, 2, "Lista:  UAT - Cenários e Passos", size=11)

    ws_capa.row_dimensions[8].height = 20
    c_cell(8, 2, "Resumo (dados de exemplo)", size=13, bold=True)

    resumo = [
        ("Total de passos",   8,   "FFFFFF"),
        ("✔  Aprovados",      4,   COR_APROVADO),
        ("✘  Reprovados",     1,   COR_REPROVADO),
        ("⚠  Bloqueados",     1,   COR_BLOQUEADO),
        ("○  Não executados", 2,   COR_NAO_EXEC),
    ]
    for i, (label, val, cor) in enumerate(resumo, start=9):
        ws_capa.row_dimensions[i].height = 20
        lc = ws_capa.cell(row=i, column=2, value=label)
        lc.font = Font(name="Calibri", size=11, color="FFFFFF")
        lc.alignment = Alignment(vertical="center")
        vc = ws_capa.cell(row=i, column=3, value=val)
        vc.font = Font(name="Calibri", size=11, bold=True,
                       color="111111" if cor != "FFFFFF" else "FFFFFF")
        vc.fill = _fill(cor)
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.border = _borda_fina()

    # ── Aba Detalhes UAT ──────────────────────────────────────────────────────
    ws = wb.create_sheet("Detalhes UAT")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    ws.merge_cells(f"A1:{get_column_letter(len(COLUNAS))}1")
    t = ws["A1"]
    t.value = f"UAT — Detalhamento de Execução   |   {date.today().strftime('%d/%m/%Y')}   |   TEMPLATE"
    t.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    t.fill = _fill(COR_TITULO)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.row_dimensions[2].height = 36
    for ci, (header, width) in enumerate(COLUNAS, start=1):
        c = ws.cell(row=2, column=ci, value=header)
        c.font = Font(name="Calibri", size=10, bold=True, color=COR_HEADER_FG)
        c.fill = _fill(COR_HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _borda_fina()
        ws.column_dimensions[get_column_letter(ci)].width = width

    cenario_atual = None
    linha_excel = 3

    for idx, row_data in enumerate(DADOS_EXEMPLO):
        id_cen, nome_cen, id_pas, desc, esp, obt, status, resp, dt, amb, ver, obs = row_data
        is_alt = (idx % 2 == 1)

        if id_cen != cenario_atual and cenario_atual is not None:
            for col in range(1, len(COLUNAS) + 1):
                ws.cell(row=linha_excel, column=col).fill = _fill("DDEEFF")
            ws.row_dimensions[linha_excel].height = 4
            linha_excel += 1
        cenario_atual = id_cen

        ws.row_dimensions[linha_excel].height = 52

        valores = [id_cen, nome_cen, id_pas, desc, esp, obt, status, resp, dt, amb, ver, obs, "—"]
        for ci, val in enumerate(valores, start=1):
            c = ws.cell(row=linha_excel, column=ci, value=val)
            c.font = Font(name="Calibri", size=10)
            c.border = _borda_fina()
            c.alignment = Alignment(vertical="top", wrap_text=True)

            if ci == 7 and status:   # coluna Status
                bg, fg = _cor_status(status)
                c.fill = _fill(bg)
                c.font = Font(name="Calibri", size=10, bold=True, color=fg)
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif ci == 13:           # coluna Evidências
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.font = Font(name="Calibri", size=10, color="999999", italic=True)
            elif is_alt:
                c.fill = _fill(COR_LINHA_ALT)

        linha_excel += 1

    ws.auto_filter.ref = f"A2:{get_column_letter(len(COLUNAS))}{linha_excel - 1}"

    # ── Aba Resumo ────────────────────────────────────────────────────────────
    ws_r = wb.create_sheet("Resumo")
    ws_r.sheet_view.showGridLines = False

    ws_r.merge_cells("A1:F1")
    h = ws_r["A1"]
    h.value = "Resumo por Cenário"
    h.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    h.fill = _fill(COR_TITULO)
    h.alignment = Alignment(horizontal="center", vertical="center")
    ws_r.row_dimensions[1].height = 30

    headers_r = ["ID Cenário", "Nome do Cenário", "Total Passos",
                 "Aprovados", "Reprovados", "Bloqueados"]
    widths_r  = [12, 38, 14, 14, 14, 14]
    for i, (h_txt, w) in enumerate(zip(headers_r, widths_r), start=1):
        c = ws_r.cell(row=2, column=i, value=h_txt)
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.fill = _fill(COR_HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _borda_fina()
        ws_r.column_dimensions[get_column_letter(i)].width = w
    ws_r.row_dimensions[2].height = 30

    cenarios_resumo = {
        "C-01": {"nome": "Login de usuário",    "T": 3, "A": 3, "R": 0, "B": 0},
        "C-02": {"nome": "Cadastro de produto", "T": 3, "A": 1, "R": 1, "B": 1},
        "C-03": {"nome": "Relatório mensal",    "T": 2, "A": 1, "R": 0, "B": 0},
    }
    for ri, (cid, info) in enumerate(cenarios_resumo.items(), start=3):
        alt = ri % 2 == 0
        vals = [cid, info["nome"], info["T"], info["A"], info["R"], info["B"]]
        for ci, v in enumerate(vals, start=1):
            c = ws_r.cell(row=ri, column=ci, value=v)
            c.font = Font(name="Calibri", size=10)
            c.border = _borda_fina()
            c.alignment = Alignment(vertical="center",
                                    horizontal="center" if ci != 2 else "left",
                                    wrap_text=True)
            if alt:
                c.fill = _fill(COR_LINHA_ALT)
            if ci == 4 and info["A"] == info["T"]:
                c.fill = _fill(COR_APROVADO)
                c.font = Font(name="Calibri", size=10, bold=True, color=COR_APROVADO_FG)
            elif ci == 5 and info["R"] > 0:
                c.fill = _fill(COR_REPROVADO)
                c.font = Font(name="Calibri", size=10, bold=True, color=COR_REPROVADO_FG)
            elif ci == 6 and info["B"] > 0:
                c.fill = _fill(COR_BLOQUEADO)
                c.font = Font(name="Calibri", size=10, bold=True, color=COR_BLOQUEADO_FG)

    ws_r.auto_filter.ref = f"A2:F{ri}"

    out = "UAT_Template.xlsx"
    wb.save(out)
    print(f"✅  Template gerado: {out}")

if __name__ == "__main__":
    gerar()
